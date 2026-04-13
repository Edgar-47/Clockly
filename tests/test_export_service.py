"""Tests for ExportService — Excel generation from attendance_sessions."""

import pytest

from app.database.connection import get_connection
from app.services.employee_service import EmployeeService
from app.services.export_service import ExportService
from app.services.time_clock_service import TimeClockService


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_employee(emp_svc: EmployeeService, *, dni: str = "EXPTEST1") -> int:
    return emp_svc.create_employee(
        first_name="Export",
        last_name="Test",
        dni=dni,
        password="clave123",
        role="employee",
    )


def _patch_exports_dir(tmp_path, monkeypatch):
    """Redirect exports to tmp_path so tests don't write to OneDrive."""
    monkeypatch.setattr("app.services.export_service.EXPORTS_DIR", tmp_path)
    monkeypatch.setattr(
        "app.services.export_service.ensure_runtime_directories", lambda: None
    )


def _insert_session(
    *,
    employee_id: int,
    clock_in: str,
    clock_out: str | None = None,
    total_seconds: int | None = None,
    is_active: bool = False,
    notes: str | None = None,
    incident_type: str | None = None,
    exit_note: str | None = None,
) -> int:
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO attendance_sessions
                (
                    user_id,
                    clock_in_time,
                    clock_out_time,
                    is_active,
                    total_seconds,
                    notes,
                    incident_type,
                    exit_note
                )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                employee_id,
                clock_in,
                clock_out,
                int(is_active),
                total_seconds,
                notes,
                incident_type,
                exit_note,
            ),
        )
        return int(cursor.lastrowid)


# ── File generation ───────────────────────────────────────────────────────────

def test_export_creates_xlsx_file(db, tmp_path, monkeypatch):
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    clk_svc = TimeClockService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    clk_svc.start_session_for_employee(emp_id)
    clk_svc.clock_out_employee(emp_id)

    path = exp_svc.export_sessions_to_excel()

    assert path.exists()
    assert path.suffix == ".xlsx"


def test_export_empty_db_does_not_create_empty_file(db, tmp_path, monkeypatch):
    """Export warns instead of generating an empty workbook."""
    _patch_exports_dir(tmp_path, monkeypatch)
    exp_svc = ExportService()

    with pytest.raises(ValueError, match="No hay sesiones"):
        exp_svc.export_sessions_to_excel()

    assert not list(tmp_path.glob("*.xlsx"))


def test_export_columns_match_sessions(db, tmp_path, monkeypatch):
    """The header row must contain the canonical session columns."""
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-13 09:00:00",
        clock_out="2026-04-13 17:00:00",
        total_seconds=28800,
    )

    path = exp_svc.export_sessions_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    assert headers == [
        "Empleado",
        "DNI",
        "Fecha entrada",
        "Hora entrada",
        "Fecha salida",
        "Hora salida",
        "Duración",
        "Estado",
        "Notas",
    ]


def test_export_data_rows_count(db, tmp_path, monkeypatch):
    """Each complete clock-in/out cycle produces exactly one data row."""
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    clk_svc = TimeClockService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    # Two complete sessions.
    clk_svc.start_session_for_employee(emp_id)
    clk_svc.clock_out_employee(emp_id)
    clk_svc.start_session_for_employee(emp_id)
    clk_svc.clock_out_employee(emp_id)

    path = exp_svc.export_sessions_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    # max_row includes the header row.
    assert ws.max_row == 3   # 1 header + 2 sessions


def test_export_workbook_includes_summary_sheet(db, tmp_path, monkeypatch):
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-13 09:00:00",
        clock_out="2026-04-13 17:00:00",
        total_seconds=28800,
    )
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-14 09:00:00",
        is_active=True,
    )

    path = exp_svc.export_sessions_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)

    assert wb.sheetnames == ["Sesiones", "Resumen"]
    summary = wb["Resumen"]
    headers = [
        summary.cell(row=1, column=c).value
        for c in range(1, summary.max_column + 1)
    ]
    assert headers == [
        "Empleado",
        "DNI",
        "Número de turnos",
        "Horas totales trabajadas",
        "Media por turno",
        "Sesiones abiertas",
        "Sesiones con incidencia",
    ]
    assert summary.cell(row=2, column=3).value == 2
    assert summary.cell(row=2, column=4).value == "08:00:00"
    assert summary.cell(row=2, column=6).value == 1


def test_export_open_session_shows_en_curso(db, tmp_path, monkeypatch):
    """A session that has not been closed must appear with 'En curso' duration."""
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    clk_svc = TimeClockService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    clk_svc.start_session_for_employee(emp_id)   # clock in, no clock out

    path = exp_svc.export_sessions_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active

    # Find the Duración column index.
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    dur_col = headers.index("Duración") + 1
    exit_col = headers.index("Hora salida") + 1
    status_col = headers.index("Estado") + 1
    duration_value = ws.cell(row=2, column=dur_col).value

    assert duration_value.startswith("En curso")
    assert ws.cell(row=2, column=exit_col).value == "Abierta"
    assert ws.cell(row=2, column=status_col).value == "abierta"


def test_export_includes_exit_note_and_incident(db, tmp_path, monkeypatch):
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    clk_svc = TimeClockService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    clk_svc.start_session_for_employee(emp_id)
    clk_svc.clock_out_employee(
        emp_id,
        exit_note="Olvido de fichaje avisado",
        incident_type="olvido",
    )

    path = exp_svc.export_sessions_to_excel()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    status_col = headers.index("Estado") + 1
    notes_col = headers.index("Notas") + 1

    assert ws.cell(row=2, column=status_col).value == "incidencia"
    assert "Olvido" in ws.cell(row=2, column=notes_col).value
    assert "Olvido de fichaje avisado" in ws.cell(row=2, column=notes_col).value


def test_build_export_filename_uses_period_and_employee_slug(db):
    exp_svc = ExportService()

    assert (
        exp_svc.build_export_filename(
            date_from="2026-04-01",
            date_to="2026-04-30",
            employee_name="Juan Pérez",
            extension="xlsx",
        )
        == "fichajes_2026-04_empleado-juan-perez.xlsx"
    )
    assert (
        exp_svc.build_export_filename(
            date_from="2026-04-02",
            date_to="2026-04-30",
            extension="pdf",
        )
        == "fichajes_2026-04-02_a_2026-04-30.pdf"
    )
    assert exp_svc.build_export_filename() == "fichajes_historico.xlsx"


def test_export_rejects_invalid_date_range(db):
    exp_svc = ExportService()

    with pytest.raises(ValueError, match="fecha desde"):
        exp_svc.validate_filters(date_from="2026-05-01", date_to="2026-04-30")


def test_export_backward_compat_alias(db, tmp_path, monkeypatch):
    """export_time_entries_to_excel() must still work (backward-compat alias)."""
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-13 09:00:00",
        clock_out="2026-04-13 17:00:00",
        total_seconds=28800,
    )

    path = exp_svc.export_time_entries_to_excel()
    assert path.exists()
    assert path.suffix == ".xlsx"


def test_export_date_filter(db, tmp_path, monkeypatch):
    """date_from / date_to filters must narrow the result set."""
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    clk_svc = TimeClockService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    clk_svc.start_session_for_employee(emp_id)
    clk_svc.clock_out_employee(emp_id)

    with pytest.raises(ValueError, match="No hay sesiones"):
        exp_svc.export_sessions_to_excel(
            date_from="2000-01-01",
            date_to="2000-12-31",
        )


def test_export_employee_filter(db, tmp_path, monkeypatch):
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc, dni="EXPTEST1")
    other_id = _make_employee(emp_svc, dni="EXPTEST2")
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-13 09:00:00",
        clock_out="2026-04-13 17:00:00",
        total_seconds=28800,
    )
    _insert_session(
        employee_id=other_id,
        clock_in="2026-04-13 10:00:00",
        clock_out="2026-04-13 18:00:00",
        total_seconds=28800,
    )

    path = exp_svc.export_sessions_to_excel(user_id=emp_id)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb["Sesiones"]

    assert ws.max_row == 2
    assert ws.cell(row=2, column=2).value == "EXPTEST1"


def test_export_sessions_to_pdf_creates_report(db, tmp_path, monkeypatch):
    pytest.importorskip("reportlab")
    _patch_exports_dir(tmp_path, monkeypatch)
    emp_svc = EmployeeService()
    exp_svc = ExportService()

    emp_id = _make_employee(emp_svc)
    _insert_session(
        employee_id=emp_id,
        clock_in="2026-04-13 09:00:00",
        clock_out="2026-04-13 17:00:00",
        total_seconds=28800,
    )

    path = exp_svc.export_sessions_to_pdf(date_from="2026-04-01", date_to="2026-04-30")

    assert path.exists()
    assert path.suffix == ".pdf"
    assert path.stat().st_size > 0
