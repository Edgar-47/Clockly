"""
ExportService - genera ficheros Excel/PDF a partir de attendance_sessions.

La fuente de verdad es attendance_sessions (no time_entries).
Cada fila del Excel representa una sesión completa con entrada, salida y
duración calculada. Las sesiones todavía abiertas se exportan indicando
que la salida está pendiente.
"""

from __future__ import annotations

import calendar
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape

from app.config import EXPORTS_DIR, ensure_runtime_directories
from app.database.attendance_session_repository import AttendanceSessionRepository
from app.services.attendance_report_service import AttendanceReportService, SessionReport


@dataclass(frozen=True)
class ExportSessionRow:
    employee_name: str
    dni: str
    entry_date: str
    entry_time: str
    exit_date: str
    exit_time: str
    duration: str
    status: str
    notes: str


@dataclass(frozen=True)
class EmployeeExportSummary:
    employee_name: str
    dni: str
    shift_count: int
    total_seconds: int
    average_seconds: int
    open_sessions: int
    incident_sessions: int


@dataclass
class _MutableEmployeeExportSummary:
    employee_name: str
    dni: str
    shift_count: int = 0
    closed_shift_count: int = 0
    total_seconds: int = 0
    open_sessions: int = 0
    incident_sessions: int = 0

    def to_summary(self) -> EmployeeExportSummary:
        average_seconds = 0
        if self.closed_shift_count:
            average_seconds = int(self.total_seconds / self.closed_shift_count)
        return EmployeeExportSummary(
            employee_name=self.employee_name,
            dni=self.dni,
            shift_count=self.shift_count,
            total_seconds=self.total_seconds,
            average_seconds=average_seconds,
            open_sessions=self.open_sessions,
            incident_sessions=self.incident_sessions,
        )


class ExportService:
    def __init__(
        self,
        attendance_session_repository: AttendanceSessionRepository | None = None,
        attendance_report_service: AttendanceReportService | None = None,
    ) -> None:
        self.attendance_session_repository = (
            attendance_session_repository or AttendanceSessionRepository()
        )
        self.attendance_report_service = (
            attendance_report_service
            or AttendanceReportService(self.attendance_session_repository)
        )

    def validate_filters(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
    ) -> tuple[str | None, str | None]:
        start, end = self._normalize_filter_dates(date_from, date_to)
        if start and end and start > end:
            raise ValueError("La fecha desde no puede ser mayor que la fecha hasta.")
        return (
            start.isoformat() if start else None,
            end.isoformat() if end else None,
        )

    def list_export_sessions(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        user_id: int | None = None,
        require_results: bool = False,
    ) -> list[SessionReport]:
        clean_date_from, clean_date_to = self.validate_filters(
            date_from=date_from,
            date_to=date_to,
        )
        reports = self.attendance_report_service.list_session_reports(
            date_from=clean_date_from,
            date_to=clean_date_to,
            user_id=user_id,
        )
        if require_results and not reports:
            raise ValueError(
                "No hay sesiones de asistencia para los filtros seleccionados."
            )
        return reports

    def export_sessions_to_excel(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        user_id: int | None = None,
        employee_name: str | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """Export attendance sessions to a two-sheet Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError(
                "Falta openpyxl. Instala dependencias con: pip install -r requirements.txt"
            ) from exc

        reports = self.list_export_sessions(
            date_from=date_from,
            date_to=date_to,
            user_id=user_id,
            require_results=True,
        )
        session_rows = self._build_session_rows(reports)
        summary_rows = self.build_summary(reports)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sesiones"

        headers = [
            "Empleado",
            "DNI",
            "Fecha entrada",
            "Hora entrada",
            "Fecha salida",
            "Hora salida",
            "Duración",
            "Estado",
            "Notas",
        ]
        sheet.append(headers)

        header_fill = PatternFill("solid", fgColor="1F538D")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in session_rows:
            sheet.append(
                [
                    row.employee_name,
                    row.dni,
                    row.entry_date,
                    row.entry_time,
                    row.exit_date,
                    row.exit_time,
                    row.duration,
                    row.status,
                    row.notes,
                ]
            )

        min_widths = {
            "A": 28,
            "B": 14,
            "C": 14,
            "D": 12,
            "E": 14,
            "F": 12,
            "G": 16,
            "H": 14,
            "I": 44,
        }
        self._format_excel_sheet(sheet, min_widths, Alignment)

        summary_sheet = workbook.create_sheet("Resumen")
        summary_headers = [
            "Empleado",
            "DNI",
            "Número de turnos",
            "Horas totales trabajadas",
            "Media por turno",
            "Sesiones abiertas",
            "Sesiones con incidencia",
        ]
        summary_sheet.append(summary_headers)
        for cell in summary_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in summary_rows:
            summary_sheet.append(
                [
                    row.employee_name,
                    row.dni,
                    row.shift_count,
                    self._format_seconds(row.total_seconds),
                    self._format_seconds(row.average_seconds),
                    row.open_sessions,
                    row.incident_sessions,
                ]
            )

        self._format_excel_sheet(
            summary_sheet,
            {
                "A": 28,
                "B": 14,
                "C": 18,
                "D": 24,
                "E": 18,
                "F": 18,
                "G": 24,
            },
            Alignment,
        )

        path = self._resolve_output_path(
            output_path=output_path,
            date_from=date_from,
            date_to=date_to,
            employee_name=employee_name or self._single_employee_name(reports),
            extension="xlsx",
        )
        workbook.save(path)
        return path

    def export_sessions_to_pdf(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        user_id: int | None = None,
        employee_name: str | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """Generate a printable PDF report from the same session dataset."""
        try:
            # ReportLab/Platypus is used because it paginates long tables cleanly.
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_RIGHT
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
            )
        except ImportError as exc:
            raise RuntimeError(
                "Falta reportlab. Instala dependencias con: pip install -r requirements.txt"
            ) from exc

        reports = self.list_export_sessions(
            date_from=date_from,
            date_to=date_to,
            user_id=user_id,
            require_results=True,
        )
        session_rows = self._build_session_rows(reports)
        summary_rows = self.build_summary(reports)
        resolved_employee_name = employee_name or self._single_employee_name(reports)
        path = self._resolve_output_path(
            output_path=output_path,
            date_from=date_from,
            date_to=date_to,
            employee_name=resolved_employee_name,
            extension="pdf",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ClockLyTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#17324D"),
            spaceAfter=8,
        )
        heading_style = ParagraphStyle(
            "ClockLyHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#17324D"),
            spaceBefore=8,
            spaceAfter=6,
        )
        normal_style = ParagraphStyle(
            "ClockLyNormal",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )
        meta_style = ParagraphStyle(
            "ClockLyMeta",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#4F5B66"),
        )
        footer_style = ParagraphStyle(
            "ClockLyFooter",
            parent=styles["BodyText"],
            alignment=TA_RIGHT,
            fontSize=7,
            textColor=colors.HexColor("#6B7280"),
        )

        def para(value: object, style=normal_style) -> Paragraph:
            return Paragraph(escape(str(value or "")), style)

        doc = SimpleDocTemplate(
            str(path),
            pagesize=landscape(A4),
            rightMargin=1.1 * cm,
            leftMargin=1.1 * cm,
            topMargin=1.0 * cm,
            bottomMargin=1.0 * cm,
        )

        elements = [
            Paragraph("Informe de fichajes", title_style),
            Paragraph(
                f"Generado: {datetime.now():%d/%m/%Y %H:%M}",
                meta_style,
            ),
            Paragraph(
                escape(
                    self.describe_filters(
                        date_from=date_from,
                        date_to=date_to,
                        employee_name=resolved_employee_name,
                    )
                ),
                meta_style,
            ),
            Spacer(1, 8),
            Paragraph("Resumen por empleado", heading_style),
        ]

        summary_data = [
            [
                "Empleado",
                "DNI",
                "Turnos",
                "Horas",
                "Media",
                "Abiertas",
                "Incidencias",
            ]
        ]
        for row in summary_rows:
            summary_data.append(
                [
                    para(row.employee_name),
                    para(row.dni),
                    row.shift_count,
                    self._format_seconds(row.total_seconds),
                    self._format_seconds(row.average_seconds),
                    row.open_sessions,
                    row.incident_sessions,
                ]
            )
        summary_table = Table(
            summary_data,
            repeatRows=1,
            colWidths=[5.0 * cm, 2.5 * cm, 1.7 * cm, 2.2 * cm, 2.2 * cm, 2.0 * cm, 2.3 * cm],
        )
        summary_table.setStyle(self._pdf_table_style(colors))
        elements.extend([summary_table, Spacer(1, 10)])

        elements.append(Paragraph("Detalle de sesiones", heading_style))
        session_data = [["Empleado", "DNI", "Entrada", "Salida", "Duración", "Estado", "Notas"]]
        for row in session_rows:
            session_data.append(
                [
                    para(row.employee_name),
                    para(row.dni),
                    para(f"{row.entry_date} {row.entry_time}".strip()),
                    para(
                        "Abierta"
                        if row.exit_time == "Abierta"
                        else f"{row.exit_date} {row.exit_time}".strip()
                    ),
                    para(row.duration),
                    para(row.status),
                    para(row.notes),
                ]
            )
        session_table = Table(
            session_data,
            repeatRows=1,
            colWidths=[4.1 * cm, 2.2 * cm, 3.0 * cm, 3.0 * cm, 2.3 * cm, 2.1 * cm, 8.4 * cm],
        )
        session_table.setStyle(self._pdf_table_style(colors))
        elements.append(session_table)

        def draw_footer(canvas, document) -> None:
            canvas.saveState()
            footer = Paragraph(f"Página {document.page}", footer_style)
            width, _ = footer.wrap(document.width, document.bottomMargin)
            footer.drawOn(
                canvas,
                document.leftMargin + document.width - width,
                0.35 * cm,
            )
            canvas.restoreState()

        doc.build(elements, onFirstPage=draw_footer, onLaterPages=draw_footer)
        return path

    def build_summary(self, reports: list[SessionReport]) -> list[EmployeeExportSummary]:
        summaries: dict[int, _MutableEmployeeExportSummary] = {}
        for report in reports:
            summary = summaries.setdefault(
                report.user_id,
                _MutableEmployeeExportSummary(
                    employee_name=report.employee_name,
                    dni=report.dni,
                ),
            )
            summary.shift_count += 1
            if report.is_active:
                summary.open_sessions += 1
            else:
                summary.closed_shift_count += 1
                summary.total_seconds += max(
                    int(report.counted_duration_seconds or 0),
                    0,
                )
            if report.has_incident:
                summary.incident_sessions += 1

        return sorted(
            (summary.to_summary() for summary in summaries.values()),
            key=lambda row: (row.employee_name.lower(), row.dni.lower()),
        )

    def build_export_filename(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        employee_name: str | None = None,
        extension: str = "xlsx",
    ) -> str:
        start, end = self._normalize_filter_dates(date_from, date_to)
        if start and end and start > end:
            raise ValueError("La fecha desde no puede ser mayor que la fecha hasta.")

        stem = self._build_period_filename_stem(start, end)
        employee_slug = self._slugify(employee_name or "")
        if employee_slug:
            stem = f"{stem}_empleado-{employee_slug}"

        clean_extension = extension.strip().lower().lstrip(".") or "xlsx"
        return f"{stem}.{clean_extension}"

    def describe_filters(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        employee_name: str | None = None,
    ) -> str:
        start, end = self._normalize_filter_dates(date_from, date_to)
        if start and end:
            period = (
                f"Periodo: {start:%Y-%m-%d}"
                if start == end
                else f"Periodo: {start:%Y-%m-%d} a {end:%Y-%m-%d}"
            )
        elif start:
            period = f"Periodo: desde {start:%Y-%m-%d}"
        elif end:
            period = f"Periodo: hasta {end:%Y-%m-%d}"
        else:
            period = "Periodo: histórico completo"

        employee = (
            f"Empleado: {employee_name.strip()}"
            if employee_name and employee_name.strip()
            else "Empleado: todos"
        )
        return f"{period} | {employee}"

    def export_time_entries_to_excel(
        self,
        *,
        date_from: str | date | None = None,
        date_to: str | date | None = None,
        employee_id: int | None = None,
        output_path: str | Path | None = None,
    ) -> Path:
        """
        Alias kept for call sites that haven't been updated yet.
        Delegates to export_sessions_to_excel using attendance_sessions.
        """
        return self.export_sessions_to_excel(
            date_from=date_from,
            date_to=date_to,
            user_id=employee_id,
            output_path=output_path,
        )

    def _build_session_rows(
        self,
        reports: list[SessionReport],
    ) -> list[ExportSessionRow]:
        return [self._build_session_row(report) for report in reports]

    def _build_session_row(self, report: SessionReport) -> ExportSessionRow:
        entry_date, entry_time = self._split_datetime(report.clock_in_time)
        exit_date = ""
        exit_time = "Abierta" if report.is_active else ""
        if report.clock_out_time:
            exit_date, exit_time = self._split_datetime(report.clock_out_time)

        duration_seconds = (
            report.display_duration_seconds
            if report.is_active
            else report.counted_duration_seconds
        )
        notes = report.notes_label
        if report.has_incident:
            notes = f"{report.incident_label}" + (f" | {notes}" if notes else "")

        return ExportSessionRow(
            employee_name=report.employee_name,
            dni=report.dni,
            entry_date=entry_date,
            entry_time=entry_time,
            exit_date=exit_date,
            exit_time=exit_time,
            duration=self._format_duration(duration_seconds, is_active=report.is_active),
            status=self._session_status(report),
            notes=notes,
        )

    def _session_status(self, report: SessionReport) -> str:
        if report.has_incident:
            return "incidencia"
        if report.is_active:
            return "abierta"
        return "cerrada"

    def _resolve_output_path(
        self,
        *,
        output_path: str | Path | None,
        date_from: str | date | None,
        date_to: str | date | None,
        employee_name: str | None,
        extension: str,
    ) -> Path:
        if output_path:
            path = Path(output_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            return path

        ensure_runtime_directories()
        filename = self.build_export_filename(
            date_from=date_from,
            date_to=date_to,
            employee_name=employee_name,
            extension=extension,
        )
        return self._unique_path(EXPORTS_DIR / filename)

    def _format_excel_sheet(self, sheet, min_widths: dict[str, int], alignment_cls) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            for cell in column:
                cell.alignment = alignment_cls(
                    vertical="top",
                    wrap_text=True,
                    horizontal="center" if cell.row == 1 else "left",
                )

        for col_letter, min_width in min_widths.items():
            max_len = max(
                (len(str(cell.value or "")) for cell in sheet[col_letter]),
                default=0,
            )
            sheet.column_dimensions[col_letter].width = max(
                min(max_len + 2, 52),
                min_width,
            )

    def _pdf_table_style(self, colors):
        from reportlab.platypus import TableStyle

        return TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )

    def _normalize_filter_dates(
        self,
        date_from: str | date | None,
        date_to: str | date | None,
    ) -> tuple[date | None, date | None]:
        return (
            self._coerce_date(date_from, "fecha desde"),
            self._coerce_date(date_to, "fecha hasta"),
        )

    def _coerce_date(self, value: str | date | None, label: str) -> date | None:
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        clean_value = str(value).strip()
        if not clean_value:
            return None
        try:
            return date.fromisoformat(clean_value)
        except ValueError as exc:
            raise ValueError(f"La {label} debe tener formato AAAA-MM-DD.") from exc

    def _build_period_filename_stem(
        self,
        date_from: date | None,
        date_to: date | None,
    ) -> str:
        if date_from and date_to:
            if self._is_full_month(date_from, date_to):
                return f"fichajes_{date_from:%Y-%m}"
            if date_from == date_to:
                return f"fichajes_{date_from:%Y-%m-%d}"
            return f"fichajes_{date_from:%Y-%m-%d}_a_{date_to:%Y-%m-%d}"
        if date_from:
            return f"fichajes_desde_{date_from:%Y-%m-%d}"
        if date_to:
            return f"fichajes_hasta_{date_to:%Y-%m-%d}"
        return "fichajes_historico"

    def _is_full_month(self, date_from: date, date_to: date) -> bool:
        last_day = calendar.monthrange(date_from.year, date_from.month)[1]
        return (
            date_from.day == 1
            and date_to.year == date_from.year
            and date_to.month == date_from.month
            and date_to.day == last_day
        )

    def _slugify(self, value: str) -> str:
        normalized = unicodedata.normalize("NFKD", value.strip().lower())
        ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
        slug = re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")
        return slug

    def _single_employee_name(self, reports: list[SessionReport]) -> str | None:
        names = {report.employee_name for report in reports if report.employee_name}
        return next(iter(names)) if len(names) == 1 else None

    def _unique_path(self, path: Path) -> Path:
        if not path.exists():
            return path

        for index in range(2, 1000):
            candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
            if not candidate.exists():
                return candidate
        return path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")

    def _split_datetime(self, value: str | None) -> tuple[str, str]:
        if not value:
            return "", ""
        try:
            dt = datetime.fromisoformat(value)
        except (TypeError, ValueError):
            return str(value), ""
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")

    def _format_duration(
        self,
        total_seconds: int | None,
        *,
        is_active: bool = False,
    ) -> str:
        if total_seconds is None:
            return "En curso" if is_active else "00:00:00"
        formatted = self._format_seconds(total_seconds)
        return f"En curso ({formatted})" if is_active else formatted

    def _format_seconds(self, total_seconds: int | None) -> str:
        seconds_value = max(int(total_seconds or 0), 0)
        hours, remainder = divmod(seconds_value, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
